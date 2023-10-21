import random
import openpyxl

class RandomNumberGenerator:
    def __init__(self, filename):
        self.filename = filename
        self.name = set()
        self.load_existing_numbers()

    def load_existing_numbers(self):
        try:
            workbook = openpyxl.load_workbook(self.filename)
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
                self.name.add(tuple(row))
        except FileNotFoundError:
            # Create a new workbook if the file doesn't exist
            self.create_new_workbook()

    def create_new_workbook(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["Generated Numbers"])
        workbook.save(self.filename)

    def save_generated_numbers(self, random_numbers):
        self.name.add(tuple(random_numbers))

        workbook = openpyxl.load_workbook(self.filename)
        worksheet = workbook.active
        worksheet.append(random_numbers)
        workbook.save(self.filename)

    def generate_and_print_random_numbers(self):
        while True:
            random_numbers = [random.randint(0, 9) for _ in range(4)]
            if tuple(random_numbers) not in self.name:
                print("Generated random numbers:", random_numbers)
                self.save_generated_numbers(random_numbers)
                break

if __name__ == "__main__":
    rng = RandomNumberGenerator("picture_name.xlsx")
    rng.generate_and_print_random_numbers()
